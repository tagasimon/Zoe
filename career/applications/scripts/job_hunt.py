#!/usr/bin/env python3
"""Generate Simon's job application packets and update the tracker.

This script does not submit applications. It prepares review-ready assets and
marks each job according to the approval-gated workflow.
"""

from __future__ import annotations

import argparse
import json
import re
import subprocess
from dataclasses import dataclass
from datetime import date, timedelta
from pathlib import Path
from typing import Iterable

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation
except ModuleNotFoundError:  # pragma: no cover - optional dependency in automation environments
    Workbook = None
    load_workbook = None
    Alignment = Font = PatternFill = DataValidation = None


ROOT = Path(__file__).resolve().parents[3]
APP_DIR = ROOT / "career" / "applications"
TRACKER = APP_DIR / "tracker.xlsx"
LOG = APP_DIR / "log.md"
JOBS_DIR = APP_DIR / "jobs"
PROFILE = APP_DIR / "profile" / "master-profile.md"
SAMPLE = APP_DIR / "samples" / "sample_jobs.json"

HEADERS = [
    "Date Found",
    "Company",
    "Role",
    "Source",
    "URL",
    "Location",
    "Job Type",
    "Fit Score",
    "Status",
    "Follow-up Date",
    "Next Action",
    "Notes",
]

STATUSES = [
    "Found",
    "Ready",
    "Drafted",
    "Needs Manual Portal",
    "Applied",
    "Followed Up",
    "Interview",
    "Rejected",
    "Withdrawn",
    "No Response",
]

ROLE_KEYWORDS = {
    "software": 3,
    "developer": 3,
    "engineer": 3,
    "full-stack": 3,
    "full stack": 3,
    "mobile": 4,
    "flutter": 5,
    "dart": 4,
    "android": 4,
    "ios": 3,
    "api": 2,
    "rest": 2,
    "python": 2,
    "django": 2,
    ".net": 2,
    "asp.net": 2,
    "kotlin": 3,
    "automation": 3,
    "rpa": 3,
    "ai": 2,
    "sql": 2,
    "power bi": 2,
    "erp": 2,
    "business central": 2,
    "sap": 1,
    "gcp": 1,
    "google cloud": 1,
}

PRODUCT_TERMS = ["product", "saas", "fintech", "marketplace", "startup", "platform"]
SENIORITY_TERMS = ["mid", "senior", "lead", "4+", "5+", "6+", "7+", "3+"]
RED_FLAGS = ["unpaid test", "unpaid assessment", "take-home project", "free trial task"]
SENSITIVE_TERMS = ["national id", "passport", "salary history", "reference", "referee", "certificate"]


@dataclass
class Job:
    company: str
    role: str
    source: str
    url: str
    location: str
    job_type: str
    apply_method: str
    apply_url: str
    email: str
    salary: str
    description: str

    @classmethod
    def from_dict(cls, raw: dict) -> "Job":
        return cls(
            company=str(raw.get("company", "")).strip(),
            role=str(raw.get("role", "")).strip(),
            source=str(raw.get("source", "")).strip() or "Unknown",
            url=str(raw.get("url", "")).strip(),
            location=str(raw.get("location", "")).strip() or "Unknown",
            job_type=str(raw.get("job_type", "")).strip() or "Unknown",
            apply_method=str(raw.get("apply_method", "")).strip().lower() or "portal",
            apply_url=str(raw.get("apply_url", "")).strip(),
            email=str(raw.get("email", "")).strip(),
            salary=str(raw.get("salary", "")).strip() or "Unknown",
            description=str(raw.get("description", "")).strip(),
        )

    @property
    def text(self) -> str:
        return f"{self.company} {self.role} {self.location} {self.job_type} {self.description}".lower()


def slugify(value: str) -> str:
    value = re.sub(r"[^a-zA-Z0-9]+", "-", value.lower()).strip("-")
    return value[:80] or "job"


def load_jobs(path: Path) -> list[Job]:
    with path.open("r", encoding="utf-8") as fh:
        data = json.load(fh)
    if not isinstance(data, list):
        raise ValueError("Input JSON must be a list of job objects.")
    return [Job.from_dict(item) for item in data if item.get("company") and item.get("role")]


def score_job(job: Job) -> tuple[int, list[str], list[str]]:
    score = 0
    reasons: list[str] = []
    risks: list[str] = []
    text = job.text

    for keyword, points in ROLE_KEYWORDS.items():
        if keyword in text:
            score += points
            reasons.append(f"+{points} {keyword}")

    if any(term in text for term in PRODUCT_TERMS):
        score += 2
        reasons.append("+2 product/company signal")

    if any(term in text for term in SENIORITY_TERMS):
        score += 1
        reasons.append("+1 mid-senior signal")

    if "remote" in text:
        score += 1
        reasons.append("+1 remote-accessible")

    if job.apply_method == "email" and job.email:
        score += 1
        reasons.append("+1 email application")

    if len(job.description) > 250:
        score += 1
        reasons.append("+1 useful JD detail")

    if job.salary.lower() == "unknown":
        risks.append("Salary unknown")

    if any(flag in text for flag in RED_FLAGS):
        score -= 10
        risks.append("Long unpaid test or speculative work flagged")

    if any(term in text for term in SENSITIVE_TERMS):
        risks.append("Sensitive information may be requested")

    if not job.description:
        score -= 10
        risks.append("Missing job description")

    return max(score, 0), reasons, risks


def choose_status(job: Job, risks: list[str]) -> str:
    if any("Sensitive information" in risk for risk in risks):
        return "Ready"
    if job.apply_method == "email" and job.email:
        return "Drafted"
    if job.apply_method == "portal" or job.apply_url:
        return "Needs Manual Portal"
    return "Ready"


def matched_skills(job: Job) -> list[str]:
    text = job.text
    skills = []
    skill_map = [
        ("Flutter / Dart mobile development", ["flutter", "dart", "mobile"]),
        ("Android / iOS app development", ["android", "ios", "kotlin"]),
        ("REST API design and integrations", ["api", "rest", "integration"]),
        ("Python / Django web development", ["python", "django"]),
        ("ASP.NET Core development", [".net", "asp.net"]),
        ("SQL databases and reporting", ["sql", "database", "reporting"]),
        ("Business automation and RPA", ["automation", "rpa", "workflow"]),
        ("ERP integrations", ["erp", "business central", "sap"]),
        ("Cloud solutions on Google Cloud", ["gcp", "google cloud", "cloud"]),
        ("Technical documentation and SOPs", ["documentation", "sop"]),
        ("Power BI and analytics", ["power bi", "analytics", "data"]),
    ]
    for label, needles in skill_map:
        if any(needle in text for needle in needles):
            skills.append(label)
    return skills or ["Full-stack software development", "Business automation", "Technical documentation"]


def professional_summary(job: Job) -> str:
    skills = matched_skills(job)[:3]
    return (
        f"Founder and Software Developer with 7+ years of experience relevant to the "
        f"{job.role} role at {job.company}. Strong background in {', '.join(skills).lower()}, "
        "with hands-on delivery across mobile apps, web systems, automation tools, ERP integrations, "
        "and business reporting. Known for turning operational requirements into practical, documented, "
        "production-ready software."
    )


def read_profile_sections() -> str:
    return PROFILE.read_text(encoding="utf-8")


def build_tailored_cv(job: Job) -> str:
    skills = matched_skills(job)
    return f"""# Kazooba Simon - Tailored CV

## Contact

- Phone: +256773383412
- Email: smkabz@gmail.com
- LinkedIn: https://www.linkedin.com/in/simon-sayz/
- Location: Kampala, Uganda
- Availability: Immediate or short notice

## Professional Summary

{professional_summary(job)}

## Core Competencies

{chr(10).join(f"- {skill}" for skill in skills)}
- Full-stack web and mobile development: Python, Django, ASP.NET Core, Flutter, Dart, Kotlin
- Databases: SQL, T-SQL, PL/SQL, Google Cloud Firestore, Oracle, MS SQL Server
- Version control: GitHub, Git

## Professional Experience

See `profile/master-profile.md` for the full source CV. This tailored draft should preserve the same employers, dates, and factual claims while reordering bullets toward:

{chr(10).join(f"- {skill}" for skill in skills)}

## Experience Highlights to Lead With

- Elastic Technologies: founder-led delivery of web, mobile, automation, product, and client systems.
- Mulwana Group: cross-platform apps using Python, Kotlin, and Flutter; ERP API integrations; Google Cloud migration.
- Nice House of Plastics: sales force automation, real-time reporting, Power BI/R analytics, predictive models.
- Culinary School Uganda: Python chatbot handling over 1,000 daily queries; server maintenance and uptime.

## Education

Bachelor of Science in Computer Science - Makerere University Kampala, 2011 to 2014

## Certifications & Training

- RESTful API Design & Development - Udemy
- Google Cloud Database Administration - Coursera
- Django Web Development - Udemy

## Achievements Relevant to Role

- Delivered ERP extensions that reduced manual data entry by over 50%.
- Developed multi-platform automation tools that cut repetitive task handling time by 70%.
- Built a sales force automation system with real-time tracking and automated reporting.
- Authored SOPs and technical documentation for developed systems.

## Referees

Available upon request.
"""


def build_cover_letter(job: Job) -> str:
    skills = matched_skills(job)[:3]
    opening = (
        f"{job.company}'s {job.role} role matches the kind of practical software work I do best: "
        f"building useful systems, integrating tools, and shipping reliable products for real users."
    )
    return f"""Dear Hiring Team,

{opening}

I bring 7+ years of experience across {', '.join(skills).lower()}. At Elastic Technologies, I lead product and client builds across web, mobile, automation, and AI-assisted workflows. At Mulwana Group of Companies, I built cross-platform applications, integrated ERP systems through APIs, and supported Google Cloud migration work. At Nice House of Plastics, I developed sales force automation and reporting systems used for operational decision-making.

I would welcome the opportunity to discuss how my background fits this role. I am available immediately or at short notice for an interview.

Yours sincerely,
Kazooba Simon
+256773383412 | smkabz@gmail.com
LinkedIn: https://www.linkedin.com/in/simon-sayz/
"""


def build_application_email(job: Job) -> str:
    skills = matched_skills(job)[:2]
    return f"""Subject: Application for {job.role} - Kazooba Simon

Dear Hiring Team,

Please find attached my CV and cover letter for the {job.role} position at {job.company}.

I am a founder and software developer with 7+ years of experience in {', '.join(skills).lower()}. I am confident my background is a strong match for this role.

I look forward to hearing from you.

Yours sincerely,
Kazooba Simon
+256773383412 | smkabz@gmail.com
LinkedIn: https://www.linkedin.com/in/simon-sayz/
"""


def build_follow_up(job: Job, application_date: date) -> str:
    return f"""Subject: Follow-up on {job.role} application - Kazooba Simon

Dear Hiring Team,

I hope you are well. I wanted to follow up on my application for the {job.role} role at {job.company}, submitted on {application_date.isoformat()}.

I remain very interested in the opportunity and would be glad to share any additional information that would help with your review.

Yours sincerely,
Kazooba Simon
+256773383412 | smkabz@gmail.com
LinkedIn: https://www.linkedin.com/in/simon-sayz/
"""


def write_packet(job: Job, score: int, status: str, reasons: list[str], risks: list[str], run_date: date) -> Path:
    folder = JOBS_DIR / run_date.isoformat() / slugify(f"{job.company}-{job.role}")
    folder.mkdir(parents=True, exist_ok=True)

    (folder / "job-description.md").write_text(
        f"""# {job.role} - {job.company}

| Field | Value |
|-------|-------|
| Source | {job.source} |
| URL | {job.url} |
| Location | {job.location} |
| Job Type | {job.job_type} |
| Apply Method | {job.apply_method or 'Unknown'} |
| Apply URL | {job.apply_url or 'N/A'} |
| Email | {job.email or 'N/A'} |
| Salary | {job.salary or 'Unknown'} |

## Description

{job.description}
""",
        encoding="utf-8",
    )

    (folder / "fit-summary.md").write_text(
        f"""# Fit Summary

| Field | Value |
|-------|-------|
| Company | {job.company} |
| Role | {job.role} |
| Fit Score | {score} |
| Status | {status} |
| Next Action | {next_action(job, status)} |

## Why It Fits

{chr(10).join(f"- {reason}" for reason in reasons) if reasons else "- No strong match signals found."}

## Risks / Flags

{chr(10).join(f"- {risk}" for risk in risks) if risks else "- No major risks flagged."}
""",
        encoding="utf-8",
    )

    (folder / "tailored-cv.md").write_text(build_tailored_cv(job), encoding="utf-8")
    (folder / "cover-letter.md").write_text(build_cover_letter(job), encoding="utf-8")
    (folder / "application-email.md").write_text(build_application_email(job), encoding="utf-8")
    (folder / "follow-up-email.md").write_text(build_follow_up(job, run_date), encoding="utf-8")
    (folder / "portal-instructions.md").write_text(
        f"""# Portal / Submission Instructions

- Status: {status}
- Apply URL: {job.apply_url or job.url or 'N/A'}
- Email: {job.email or 'N/A'}
- Do not submit until Simon reviews the packet.
- If the portal asks for references, salary history, ID, certificates, or extra documents, stop and ask Simon.
- After Simon confirms submission, update tracker status to `Applied` and set follow-up date to {(run_date + timedelta(days=7)).isoformat()}.
""",
        encoding="utf-8",
    )
    return folder


def next_action(job: Job, status: str) -> str:
    if status == "Drafted":
        return "Review generated email, attach CV/cover letter, then send if approved"
    if status == "Needs Manual Portal":
        return "Review packet and complete portal manually; stop for sensitive fields"
    return "Review fit and decide whether to proceed"


def notes(job: Job, risks: list[str], packet: Path) -> str:
    parts = []
    if job.salary.lower() == "unknown":
        parts.append("Pay unknown")
    parts.extend(risk for risk in risks if risk != "Salary unknown")
    parts.append(f"Packet: {packet.relative_to(ROOT)}")
    return "; ".join(dict.fromkeys(parts))


def ensure_tracker() -> None:
    if Workbook is None:
        return
    if TRACKER.exists():
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"
    ws.append(HEADERS)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = [14, 28, 34, 18, 48, 20, 16, 12, 22, 16, 36, 60]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:L1"
    validation = DataValidation(type="list", formula1=f'"{",".join(STATUSES)}"', allow_blank=False)
    ws.add_data_validation(validation)
    validation.add("I2:I500")

    summary = wb.create_sheet("Summary")
    summary["A1"] = "Career Application Tracker"
    summary["A1"].font = Font(size=16, bold=True, color="1F4E78")
    summary["A3"] = "Metric"
    summary["B3"] = "Value"
    summary["A4"] = "Total Rows"
    summary["B4"] = '=COUNTA(Applications!B:B)-1'
    summary["A5"] = "Applied"
    summary["B5"] = '=COUNTIF(Applications!I:I,"Applied")'
    summary["A6"] = "Interviews"
    summary["B6"] = '=COUNTIF(Applications!I:I,"Interview")'
    summary["A7"] = "Needs Manual Portal"
    summary["B7"] = '=COUNTIF(Applications!I:I,"Needs Manual Portal")'
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 16
    wb.save(TRACKER)


def existing_keys(ws) -> set[tuple[str, str, str]]:
    keys = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[1] and not row[2] and not row[4]:
            continue
        keys.add((str(row[1] or ""), str(row[2] or ""), str(row[4] or "")))
    return keys


def update_tracker(rows: Iterable[list]) -> None:
    if load_workbook is None:
        return
    ensure_tracker()
    wb = load_workbook(TRACKER)
    ws = wb["Applications"]
    keys = existing_keys(ws)
    for row in rows:
        key = (row[1], row[2], row[4])
        if key in keys:
            continue
        ws.append(row)
        keys.add(key)
    wb.save(TRACKER)


def ensure_log() -> None:
    if LOG.exists():
        return
    LOG.write_text(
        "# Job Applications Log\n\n"
        "| Date | Company | Role | Source URL | Status | Notes |\n"
        "|------|---------|------|------------|--------|-------|\n",
        encoding="utf-8",
    )


def append_log(rows: Iterable[list]) -> None:
    ensure_log()
    existing = LOG.read_text(encoding="utf-8")
    additions = []
    for row in rows:
        line = f"| {row[0]} | {row[1]} | {row[2]} | {row[4]} | {row[8]} | {row[11]} |\n"
        if line not in existing:
            additions.append(line)
    if additions:
        marker = "\n---\n"
        if marker in existing:
            before, after = existing.split(marker, 1)
            LOG.write_text(before.rstrip() + "\n" + "".join(additions) + marker + after, encoding="utf-8")
        else:
            with LOG.open("a", encoding="utf-8") as fh:
                fh.writelines(additions)


def create_gmail_draft(job: Job, folder: Path) -> bool:
    """Best-effort Gmail draft creation. Never sends email."""
    if not job.email:
        return False
    email_text = (folder / "application-email.md").read_text(encoding="utf-8")
    subject_match = re.search(r"^Subject:\s*(.+)$", email_text, re.MULTILINE)
    subject = subject_match.group(1).strip() if subject_match else f"Application for {job.role} - Kazooba Simon"
    body = re.sub(r"^Subject:\s*.+\n\n?", "", email_text, count=1, flags=re.MULTILINE)
    command = [
        "gws",
        "gmail",
        "draft",
        "--to",
        job.email,
        "--subject",
        subject,
        "--body",
        body,
    ]
    (folder / "gmail-draft-command.txt").write_text(" ".join(command), encoding="utf-8")
    try:
        completed = subprocess.run(command, check=False, capture_output=True, text=True, timeout=30)
    except (FileNotFoundError, subprocess.TimeoutExpired):
        return False
    (folder / "gmail-draft-result.txt").write_text(
        f"returncode={completed.returncode}\n\nSTDOUT:\n{completed.stdout}\n\nSTDERR:\n{completed.stderr}",
        encoding="utf-8",
    )
    return completed.returncode == 0


def run(input_path: Path, limit: int, min_score: int, gmail_drafts: bool) -> list[tuple[Job, int, str, Path]]:
    run_date = date.today()
    jobs = load_jobs(input_path)
    ranked = []
    for job in jobs:
        score, reasons, risks = score_job(job)
        if score < min_score:
            continue
        ranked.append((score, job, reasons, risks))
    ranked.sort(key=lambda item: item[0], reverse=True)

    tracker_rows = []
    results = []
    for score, job, reasons, risks in ranked[:limit]:
        status = choose_status(job, risks)
        packet = write_packet(job, score, status, reasons, risks, run_date)
        if gmail_drafts and status == "Drafted":
            created = create_gmail_draft(job, packet)
            if not created:
                status = "Ready"
        follow_up = "" if status != "Applied" else (run_date + timedelta(days=7)).isoformat()
        row = [
            run_date.isoformat(),
            job.company,
            job.role,
            job.source,
            job.url,
            job.location,
            job.job_type,
            score,
            status,
            follow_up,
            next_action(job, status),
            notes(job, risks, packet),
        ]
        tracker_rows.append(row)
        results.append((job, score, status, packet))

    update_tracker(tracker_rows)
    append_log(tracker_rows)
    return results


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Prepare Simon's job application packets.")
    parser.add_argument("--input", type=Path, default=SAMPLE, help="JSON file containing job listings.")
    parser.add_argument("--limit", type=int, default=10, help="Maximum number of packets to generate.")
    parser.add_argument("--min-score", type=int, default=5, help="Minimum fit score required.")
    parser.add_argument("--create-gmail-drafts", action="store_true", help="Create Gmail drafts with gws when email is available. Never sends.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if not args.input.exists():
        raise SystemExit(f"Input file not found: {args.input}")
    if not PROFILE.exists():
        raise SystemExit(f"Master profile not found: {PROFILE}")
    results = run(args.input, args.limit, args.min_score, args.create_gmail_drafts)
    print(f"Prepared {len(results)} application packet(s).")
    for job, score, status, packet in results:
        print(f"- {job.company} | {job.role} | score {score} | {status} | {packet}")


if __name__ == "__main__":
    main()
